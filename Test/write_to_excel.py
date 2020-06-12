from openpyxl import Workbook
from openpyxl import load_workbook





class Excel_operation():
    def create_excel(self):
        headers=['Country','City','Date','Night', 'Adult', 'Hotel Name', 'Hotel Star Rating', 'Room Name', 'Beds & Amenities', 'Avg. Price Per Room / Night']
        workbook_name ='round_trip.xlsx'
        wb = Workbook()
        page = wb.active
        page.title = 'hotels'
        page.append(headers)
        wb.save(filename=workbook_name)

    def append_to_execl(self,data):

        workbook_name = 'round_trip.xlsx'
        wb = load_workbook(workbook_name)
        page = wb.active
        page.append(data)
        wb.save(filename=workbook_name)

a=Excel_operation()

#a.write_to_excel()
# data1=[' China', 'Beijing', '14th Mar 2021', 2, 5, 'Xinyue Apartment Hotel', 3, 'Duplex Room', 'No meals included', '219.30']
# data2=[' China', 'Beijing', '14th Mar 2021', 2, 5, 'Xinyue Apartment Hotel', 3, 'Duplex Room', 'No meals included', '219.30']
# a.append_to_execl(data1)
# a.append_to_execl(data2)


